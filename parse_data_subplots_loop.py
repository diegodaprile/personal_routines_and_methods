# -*- coding: utf-8 -*-
"""
Created on Tue Dec 19 19:37:27 2017

@author: User
"""

# -*- coding: utf-8 -*-
"""
Created on Tue Dec 19 11:56:13 2017

@author: User
"""

from openpyxl import load_workbook
import matplotlib.pyplot as plt
 
def parser_to_dictionary(sheet):
    columns = []
    for row in sheet.rows:
        if columns:
                [columns[i].append(ref.value) for i, ref in enumerate(row)]
        else:
            columns = [[ref.value] for ref in row]
    result = {x[0] : x[1:] for x in columns}
    return result


sheet_ranges1 = load_workbook(filename = 'dataset_example_1.xlsx')['Sheet1']
sheet_ranges2 = load_workbook(filename = 'dataset_example_2.xlsx')['Sheet1']
data = {**parser_to_dictionary(sheet_ranges1), **parser_to_dictionary(sheet_ranges2)}

# variables of interest
varOfInterest = ["young_aged_men_percent_c_", "mid_aged_men_percent_c_", "next_to_pen_men_percent_c_", \
                 "young_aged_women_percent_c_", "mid_aged_women_percent_c_", "next_to_pen_women_percent_c_", "gdp_per_capita_ppp_usd_c_"]

countrySelectionDict = {21:"Ireland", 27:"Luxembourg", 33:"Norway", 43:"Switzerland"}

x = data['time']


for key, value in countrySelectionDict.items():
    
    f, axs = plt.subplots(3,2, sharex='col', figsize=(10,8))
    axs = axs.ravel()
    
    for i in range(6):
        axs[i].plot(x, data[varOfInterest[i]+str(key)], linewidth=2, c='#F4A460')
        axs[i].set_xlabel("Time")
        axs[i].set_ylabel("Pop, perc") 
        axs[i].set_title(varOfInterest[i]+countrySelectionDict[key])
        axs[i].set_axis_bgcolor("#FFFF66")
    plt.tight_layout()
    #plt.savefig('example.svg', bbox_inches='tight')  
    plt.show()

